
from django.shortcuts import render, redirect, get_object_or_404
from .models import Tramite, Etapa, Usuario
from .forms import TramiteForm
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponseForbidden
from .forms import FiltroTramitesForm
from .forms import MoverTramiteForm
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
from .forms import ImportTramiteForm
from django.utils import timezone
from datetime import timedelta
from django.core.exceptions import ValidationError
from .forms import CustomUserCreationForm

# importaciones para el login

from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import login, authenticate
from django.shortcuts import render, redirect
from django.contrib import messages

# importaciones para el logup

from django.contrib.auth import logout




# método para válidar el rol de un usuario
#---------------------------------------------------------------------------------------------------------------------------------------------

def check_role(user, role):
    if user.rol != role:
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
    return None

# Vista de inicio 
#---------------------------------------------------------------------------------------------------------------------------------------------

@login_required
def home(request):
    return render(request, 'app_catastro_inmuebles_01/home.html')

# Vista para crear un nuevo trámite 
#---------------------------------------------------------------------------------------------------------------------------------------------

@login_required
def crear_tramite(request):
    if request.method == 'POST':
        form = TramiteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('bandeja_inspeccion')  # Redirige a la bandeja de inspección después de crear el trámite
    else:
        form = TramiteForm()
    
    return render(request, 'app_catastro_inmuebles_01/crear_tramite.html', {'form': form})

# Vista para mover un trámite a otra etapa
#---------------------------------------------------------------------------------------------------------------------------------------------

# Vista para cambiar la etapa de un trámite
#---------------------------------------------------------------------------------------------------------------------------------------------


# Vista para mostrar la bandeja de recepción (ok)
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_recepcion(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """
    # Obtener todos los trámites en la etapa de "Recepción"
    tramites = Tramite.objects.filter(etapa__nombre="Recepción")
    inspectores = Usuario.objects.filter(rol='inspector')  # Obtener todos los inspectores disponibles
    
    # Obtener todas las etapas disponibles
    etapas = Etapa.objects.all()  
    
    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'inspectores': inspectores,
        'tramites': tramites_ordenados,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }


    return render(request, 'app_catastro_inmuebles_01/bandeja_recepcion.html', context)


# Vista para mostrar la bandeja de inspección (ok)
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_inspeccion(request):


    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    tramites = Tramite.objects.filter(etapa__nombre="Inspección")	
    etapas = Etapa.objects.all()
    resultores = Usuario.objects.filter(rol='resultor')  # Obtener todos los resultores disponibles
    lideres_proyecto = Usuario.objects.filter(rol='lider_proyecto')  # Obtener todos los líderes de proyecto disponibles
    
    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'resultores': resultores,
        'lideres_proyecto': lideres_proyecto,
        'tramites': tramites_ordenados,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }
   
    return render(request, 'app_catastro_inmuebles_01/bandeja_inspeccion.html', context)

# Vista para mostrar la bandeja de resolucion
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_resolucion(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    tramites = Tramite.objects.filter(etapa__nombre="Resolución")	
    etapas = Etapa.objects.all()

    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'tramites': tramites_ordenados,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }
   


    return render(request, 'app_catastro_inmuebles_01/bandeja_resolucion.html', context)


# Vista para mostrar la bandeja de resolucion Proyectos
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_resolucion_proyectos(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    tramites = Tramite.objects.filter(etapa__nombre="Resolución Proyecto")	
    etapas = Etapa.objects.all()

    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'tramites': tramites_ordenados,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }
   


    return render(request, 'app_catastro_inmuebles_01/bandeja_resolucion_proyecto.html', context)



# Vista para mostrar la bandeja de resolucion
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_cola_notificacion(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    tramites = Tramite.objects.filter(etapa__nombre="Cola Notificación")	
    etapas = Etapa.objects.all()  
    
    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'tramites': tramites_ordenados,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }   
 

    return render(request, 'app_catastro_inmuebles_01/bandeja_cola_notificacion.html', context)


# Vista para mostrar la bandeja de finalizados
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def bandeja_finalizados(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    # Obtener todos los trámites en la etapa de Recepción
    tramites = Tramite.objects.filter(etapa__nombre="Finalizado")	

    etapas = Etapa.objects.all()  # Obtener todas las etapas disponibles
    
    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    tramites_totales = len(tramites)

    context = {
        'etapas': etapas,
        'tramites': tramites,
        'tramites_totales': tramites_totales,
        'form': form,
    }
    return render(request, 'app_catastro_inmuebles_01/bandeja_finalizados.html', context)


from datetime import timedelta
from django.utils import timezone

@login_required
def bandeja_inspeccion_proyectos(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    # Obtener todos los trámites en la etapa de "Inspección Proyecto"
    tramites = Tramite.objects.filter(etapa__nombre="Inspección Proyecto")
    
    # Obtener todas las etapas disponibles
    etapas = Etapa.objects.all()  

    sub_lideres_proyecto = Usuario.objects.filter(rol='sub_lider_proyecto')  # Obtener todos los sub-líderes de proyecto disponibles

    resultores_proyectos = Usuario.objects.filter(rol='resultor_proyecto')  # Obtener todos los resultores de proyecto disponibles
    
    # Filtrar los trámites según el formulario de filtros
    if request.method == 'GET':
        form = FiltroTramitesForm(request.GET)
        if form.is_valid():
            tramites = form.filter_tramites(tramites)
    else:
        form = FiltroTramitesForm()

    # Definir las fechas de corte
    today = timezone.now()
    warning_period = timedelta(days=5)  # Definir 5 días antes del vencimiento como "por vencer"

    # Clasificar los trámites en tres categorías
    tramites_vencidos = []
    tramites_por_vencer = []
    tramites_ok = []

    for tramite in tramites:
        if tramite.fecha_vencimiento < today:
            tramite.estado = 'vencido'
            tramites_vencidos.append(tramite)
        elif tramite.fecha_vencimiento <= today + warning_period:
            tramite.estado = 'por_vencer'
            tramites_por_vencer.append(tramite)
        else:
            tramite.estado = 'ok'
            tramites_ok.append(tramite)
    
    # Contar los trámites en cada estado
    tramites_vencidos_count = len(tramites_vencidos)
    tramites_por_vencer_count = len(tramites_por_vencer)
    tramites_ok_count = len(tramites_ok)
    tramites_totales = tramites_vencidos_count + tramites_por_vencer_count + tramites_ok_count

    # Concatenar los trámites ordenados: vencidos, por vencer y ok
    tramites_ordenados = tramites_vencidos + tramites_por_vencer + tramites_ok
    
    # Pasar los datos al contexto
    context = {
        'etapas': etapas,
        'tramites': tramites_ordenados,
        'sub_lideres_proyecto': sub_lideres_proyecto,
        'resultores_proyectos': resultores_proyectos,
        'tramites_vencidos_count': tramites_vencidos_count,
        'tramites_por_vencer_count': tramites_por_vencer_count,
        'tramites_ok_count': tramites_ok_count,
        'tramites_totales': tramites_totales,
        'form': form,
    }
    
    return render(request, 'app_catastro_inmuebles_01/bandeja_inspeccion_proyecto.html', context)


# Vista para cambiar de etapa
#---------------------------------------------------------------------------------------------------------------------------------------------

@login_required
def cambiar_etapa(request, tramite_id):
    # Obtener el trámite correspondiente
    tramite = get_object_or_404(Tramite, id=tramite_id)

    if request.method == 'POST':
        # Obtener los datos del formulario desde el POST
        nueva_etapa_id = request.POST.get('etapa')
        inpector_id = request.POST.get('inspector')
        resultor_responsable_id = request.POST.get('resultor')
        lider_proyecto_id = request.POST.get('lider_proyecto')
        sub_lider_proyecto_id = request.POST.get('sub_lider_proyecto')
        resultor_proyecto_id = request.POST.get('resultor_proyecto')



        # Validar que la etapa seleccionada existe
        if nueva_etapa_id:
            nueva_etapa = get_object_or_404(Etapa, id=nueva_etapa_id)
            tramite.etapa = nueva_etapa

        # Si se ha proporcionado un inspector, asignarlo
        if inpector_id:
            inspector = get_object_or_404(Usuario, id=inpector_id)
            tramite.inspector_responsable = inspector

        # Si se ha proporcionado un resultor, asignarlo
        if resultor_responsable_id:
            resultor_responsable = get_object_or_404(Usuario, id=resultor_responsable_id)
            tramite.resultor_responsable = resultor_responsable

        # Si se ha proporcionado un líder de proyecto, asignarlo
        if lider_proyecto_id:
            lider_proyecto = get_object_or_404(Usuario, id=lider_proyecto_id)
            tramite.lider_proyecto = lider_proyecto

        # Si se ha proporcionado un sub-líder de proyecto, asignarlo
        if sub_lider_proyecto_id:
            sub_lider_proyecto = get_object_or_404(Usuario, id=sub_lider_proyecto_id)
            tramite.sub_lider_proyecto = sub_lider_proyecto
        
        # Si se ha proporcionado un resultor de proyecto, asignarlo
        if resultor_proyecto_id:
            resultor_proyecto = get_object_or_404(Usuario, id=resultor_proyecto_id)
            tramite.resultor_proyecto = resultor_proyecto

        # Guardar el trámite con los nuevos valores
        tramite.save()

        # Mostrar un mensaje de éxito
        messages.success(request, f'El trámite {tramite.no_oficio} se ha actualizado correctamente.')

        # Redirigir al usuario a la página de la bandeja
        return redirect(request.META.get('HTTP_REFERER', 'home'))

    # Si el método no es POST, mostrar un mensaje de error
    messages.error(request, 'Método no permitido.')
    return redirect(request.META.get('HTTP_REFERER', 'home'))


# vista para importar trámites desde excel
#---------------------------------------------------------------------------------------------------------------------------------------------
@login_required
def import_tramites(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """


    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Abrir el archivo Excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error al leer el archivo Excel: {e}")
            return redirect('import_tramites')
        
        # Procesar las filas del archivo
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)): # Asumimos que la primera fila es encabezado
            try:
                no_oficio = row[0]
                fecha_recepcion = row[1]
                fecha_recepcion = timezone.make_aware(fecha_recepcion)
                fecha_vencimiento = row[2]
                fecha_vencimiento = timezone.make_aware(fecha_vencimiento)
                prioridad = row[3]
                tramite_tipo = row[4]
                comentarios = row[5]
                etapa_id = row[6]
                inspector_responsable_id = row[7]
                receptor_responsable_id = row[8]
                resultor_responsable_id = row[9]
                lider_proyecto_id = row[10]
                sub_lider_proyecto_id = row[11]
                resultor_proyecto_id = row[12]

                # Validar que todos los campos sean correctos
                if not (no_oficio and fecha_recepcion and fecha_vencimiento and prioridad and tramite_tipo and comentarios  and etapa_id and inspector_responsable_id and receptor_responsable_id):
                    raise ValidationError(f"Algunos campos del trámite {no_oficio}  están vacíos.")
                
                # Validar que las fechas sean correctas
                if fecha_recepcion > fecha_vencimiento:
                    raise ValidationError(f"La fecha de recepción para el trámite {no_oficio} es mayor a la fecha de vencimiento.")

                # Validar que la etapa exista
                etapa = Etapa.objects.filter(id=etapa_id).first()
                if not etapa:
                    raise ValidationError(f"La etapa con id {etapa_id} no existe.")
                
                # Validar los responsables (inspector, receptor, resultor, etc.)
                inspector_responsable = Usuario.objects.filter(id=inspector_responsable_id).first()
                receptor_responsable = Usuario.objects.filter(id=receptor_responsable_id).first()
                resultor_responsable = Usuario.objects.filter(id=resultor_responsable_id).first()
                lider_proyecto = Usuario.objects.filter(id=lider_proyecto_id).first()
                sub_lider_proyecto = Usuario.objects.filter(id=sub_lider_proyecto_id).first()
                resultor_proyecto = Usuario.objects.filter(id=resultor_proyecto_id).first()

                if not (inspector_responsable and receptor_responsable):
                    raise ValidationError(f"Faltan responsables para el trámite {no_oficio}.")

                # Crear el trámite
                Tramite.objects.create(
                    no_oficio=no_oficio,
                    fecha_recepcion=fecha_recepcion,
                    fecha_vencimiento=fecha_vencimiento,
                    prioridad=prioridad,
                    tramite_tipo=tramite_tipo,
                    comentarios=comentarios,
                    etapa=etapa,
                    inspector_responsable=inspector_responsable,
                    receptor_responsable=receptor_responsable,
                    resultor_responsable=resultor_responsable,
                    lider_proyecto=lider_proyecto,
                    sub_lider_proyecto=sub_lider_proyecto,
                    resultor_proyecto=resultor_proyecto

                )

            except ValidationError as e:
                messages.error(request, f"Error en la fila {idx + 2}: {e} - Registro no importado.")
                continue
            except Exception as e:
                messages.error(request, f"Error en la fila {idx + 2}: {e} - Registro no importado.")
                continue

        messages.success(request, "Los trámites se importaron correctamente.")
        return redirect('import_tramites')

    return render(request, 'app_catastro_inmuebles_01/import_tramites.html')


@login_required
def export_tramites(request):

    """
    # Lista de roles permitidos
    roles_permitidos = ['resultor', 'inspector', 'receptor', 'lider_proyecto', 'sub_lider_proyecto']

    # Verificar si el usuario tiene uno de los roles permitidos
    role_check = check_role(request.user, roles_permitidos)
    if role_check:
        return role_check
    """

    # Obtener todos los trámites
    tramites = Tramite.objects.all()
    
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Trámites"
    
    # Escribir los encabezados de las columnas
    headers = [
        "No. Oficio", "Fecha Recepción", "Fecha Vencimiento", "Prioridad", "Tipo Trámite",
        "Comentarios", "Etapa", "Receptor Responsable", "Inspector Responsable",  
        "Resultor Responsable", "Líder Proyecto", "Sub Líder Proyecto", "Resultor Proyecto" "Estado"
    ]
    
    sheet.append(headers)
    
    # Escribir los datos de cada trámite
    for tramite in tramites:
        data = [
            tramite.no_oficio,
            tramite.fecha_recepcion.strftime("%Y-%m-%d %H:%M:%S"),
            tramite.fecha_vencimiento.strftime("%Y-%m-%d %H:%M:%S"),
            tramite.prioridad,
            tramite.tramite_tipo,
            tramite.comentarios if tramite.comentarios else '',
            tramite.etapa.nombre,
            tramite.receptor_responsable.username if tramite.receptor_responsable else '',
            tramite.inspector_responsable.username if tramite.inspector_responsable else '',
            tramite.resultor_responsable.username if tramite.resultor_responsable else '',
            tramite.lider_proyecto.username if tramite.lider_proyecto else '',
            tramite.sub_lider_proyecto.username if tramite.sub_lider_proyecto else '',
            tramite.resultor_proyecto.username if tramite.resultor_proyecto else '',

            tramite.get_estado()  # Llamar al método que determina el estado (vencido, por vencer, ok)
        ]
        sheet.append(data)
    
    # Crear una respuesta HTTP que permita descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tramites.xlsx"'
    
    # Guardar el libro en la respuesta HTTP
    wb.save(response)
    
    return response

def login_amsac(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, "Credenciales incorrectas")
        else:
            print(form.errors)  # Imprimir errores del formulario
            messages.error(request, "Credenciales incorrectas.")
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})

# Vista para el registro de usuarios

def signup(request):

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro exitoso. Ahora puedes iniciar sesión.')
            return redirect('login_amsac')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = CustomUserCreationForm()
    return render(request, 'app_catastro_inmuebles_01/signup.html', {'form': form})


# Vista para cerrar sesión


def logout_amsac(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('login_amsac')